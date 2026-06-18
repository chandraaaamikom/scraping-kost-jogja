from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

RAW_FILE = Path("data/raw/postingan_kost.txt")

OUTPUT_ALL_CSV = Path("data/clean/data_kost_jogja.csv")
OUTPUT_ALL_EXCEL = Path("data/clean/data_kost_jogja.xlsx")

OUTPUT_FILTER_CSV = Path("data/clean/data_kost_jogja_filter.csv")
OUTPUT_FILTER_EXCEL = Path("data/clean/data_kost_jogja_filter.xlsx")


def baca_data_mentah():
    if not RAW_FILE.exists():
        print("File postingan_kost.txt tidak ditemukan.")
        return []

    teks = RAW_FILE.read_text(encoding="utf-8")
    postingan = teks.split("---")

    hasil = []
    for post in postingan:
        post = post.strip()
        if post:
            hasil.append(post)

    return hasil


def ambil_judul(teks):
    baris = teks.strip().splitlines()
    for item in baris:
        item = item.strip()
        if item:
            return item[:100]
    return None


def cek_status_postingan(teks):
    """
    Membedakan postingan penawaran kost dan postingan orang mencari kost.
    """
    teks_lower = teks.lower().strip()

    kata_pencarian = [
        "cari kost",
        "cari kos",
        "mencari kost",
        "mencari kos",
        "harga maksimal",
        "sertakan foto",
    ]

    for kata in kata_pencarian:
        if kata in teks_lower:
            return "Pencarian"

    if teks_lower.startswith("info kos") and "budget" in teks_lower:
        return "Pencarian"

    if teks_lower.startswith("info kost") and "budget" in teks_lower:
        return "Pencarian"

    return "Penawaran"


def konversi_harga(angka, satuan):
    """
    Mengubah harga menjadi satuan ribu rupiah.
    Contoh:
    - 650rb  -> 650
    - 700k   -> 700
    - 1 juta -> 1000
    """
    angka = angka.replace(",", ".")
    angka = float(angka)

    satuan = satuan.lower()

    if satuan in ["rb", "ribu", "k"]:
        return int(angka)

    if satuan in ["jt", "juta"]:
        return int(angka * 1000)

    return None


def ambil_semua_harga(teks):
    """
    Membaca semua harga dari teks.
    Mendukung:
    - 650rb
    - 700rb/bulan
    - 900k/bulan
    - 1 juta
    - 1,2 juta
    """
    teks_lower = teks.lower()
    pola = r"(\d+(?:[,.]\d+)?)\s*(rb|ribu|k|jt|juta)"

    hasil = []
    for match in re.finditer(pola, teks_lower):
        angka = match.group(1)
        satuan = match.group(2)
        harga = konversi_harga(angka, satuan)

        if harga:
            hasil.append(harga)

    return hasil


def ambil_area(teks):
    """
    Area diprioritaskan dari lokasi utama, bukan dari bagian akses.
    Contoh:
    Pugeran Maguwoharjo ... akses 3 menit dari Seturan
    harus terbaca Maguwoharjo, bukan Seturan.
    """
    teks_lower = teks.lower()

    # Ambil bagian awal saja agar area akses tidak terlalu dominan
    bagian_awal = teks_lower[:500]

    daftar_area_prioritas = [
        ("maguwoharjo", "Maguwoharjo"),
        ("stadion maguwoharjo", "Maguwoharjo"),
        ("sambilegi", "Maguwoharjo"),
        ("condongcatur", "Condongcatur"),
        ("condong catur", "Condongcatur"),
        ("concat", "Condongcatur"),
        ("catur tunggal", "Catur Tunggal"),
        ("jakal", "Jalan Kaliurang"),
        ("jalan kaliurang", "Jalan Kaliurang"),
        ("jl kaliurang", "Jalan Kaliurang"),
        ("babarsari", "Babarsari"),
        ("seturan", "Seturan"),
        ("pogung", "Pogung"),
        ("gejayan", "Gejayan"),
        ("nologaten", "Nologaten"),
        ("ambarukmo", "Ambarukmo"),
        ("pugeran", "Pugeran"),
        ("ngemplak", "Ngemplak"),
        ("sleman", "Sleman"),
        ("depok", "Depok"),
        ("ugm", "UGM"),
    ]

    for kata, nama_area in daftar_area_prioritas:
        if kata in bagian_awal:
            return nama_area

    for kata, nama_area in daftar_area_prioritas:
        if kata in teks_lower:
            return nama_area

    return None


def ambil_tipe_kost(teks):
    teks_lower = teks.lower()

    ada_putra = "putra" in teks_lower or "cowok" in teks_lower or "laki" in teks_lower
    ada_putri = "putri" in teks_lower or "cewek" in teks_lower or "perempuan" in teks_lower
    ada_campur = "campur" in teks_lower
    ada_pasutri = "pasutri" in teks_lower

    if ada_putra and ada_campur:
        return "Putra/Campur"

    if ada_putra and ada_putri:
        return "Putra/Putri"

    if ada_campur:
        return "Campur"

    if ada_putra:
        return "Putra"

    if ada_putri:
        return "Putri"

    if ada_pasutri:
        return "Pasutri"

    return None


def ambil_status_kamar_mandi(teks):
    teks_lower = teks.lower()

    if "kamar mandi dalam" in teks_lower or "km dalam" in teks_lower or "dapur didalam" in teks_lower:
        return "Dalam"

    if (
        "kamar mandi luar" in teks_lower
        or "km luar" in teks_lower
        or "kamar mandi diluar" in teks_lower
        or "kamar mandi di luar" in teks_lower
    ):
        return "Luar"

    return None


def ambil_fasilitas(teks):
    teks_lower = teks.lower()

    daftar_fasilitas = {
        "wifi": "WiFi",
        "wi-fi": "WiFi",
        "free wifi": "WiFi",
        "ac": "AC",
        "kasur": "Kasur",
        "lemari": "Lemari",
        "kamar mandi dalam": "Kamar Mandi Dalam",
        "km dalam": "Kamar Mandi Dalam",
        "kamar mandi luar": "Kamar Mandi Luar",
        "km luar": "Kamar Mandi Luar",
        "parkir motor": "Parkir Motor",
        "parkiran": "Parkir",
        "parkir": "Parkir",
        "dapur umum": "Dapur Umum",
        "dapur bersama": "Dapur Bersama",
        "dapur didalam": "Dapur Dalam",
        "listrik": "Listrik",
        "air": "Air",
        "sampah": "Sampah",
        "akses 24 jam": "Akses 24 Jam",
        "bebas jam malam": "Bebas Jam Malam",
    }

    fasilitas = []

    for kata, nama in daftar_fasilitas.items():
        if kata in teks_lower:
            fasilitas.append(nama)

    return list(dict.fromkeys(fasilitas))


def baris_valid_harga(line):
    """
    Menentukan apakah baris harga layak dianggap harga bulanan/kamar.
    Menghindari paket seperti 2550k/3 bulan, 4800k/6 bulan, 9000k/1 tahun.
    """
    line_lower = line.lower()

    if "tahun" in line_lower:
        return False

    if re.search(r"/\s*\d+\s*bulan", line_lower):
        return False

    if re.search(r"\d+\s*bulan", line_lower) and not re.search(r"/\s*(bulan|bln)", line_lower):
        return False

    return True


def buat_dataframe(data_postingan):
    hasil_data = []
    nomor_id = 1

    for nomor_post, post in enumerate(data_postingan, start=1):
        status = cek_status_postingan(post)
        judul = ambil_judul(post)
        area = ambil_area(post)
        tipe_kost = ambil_tipe_kost(post)
        fasilitas = ambil_fasilitas(post)

        lines = post.splitlines()
        ada_baris_harga = False

        for line in lines:
            harga_list = ambil_semua_harga(line)

            if not harga_list:
                continue

            if not baris_valid_harga(line):
                continue

            ada_baris_harga = True

            for harga in harga_list:
                km_line = ambil_status_kamar_mandi(line)

                if not km_line:
                    km_line = ambil_status_kamar_mandi(post)

                data = {
                    "id": nomor_id,
                    "nomor_postingan": nomor_post,
                    "status": status,
                    "judul": judul,
                    "area": area,
                    "tipe_kost": tipe_kost,
                    "harga": harga,
                    "kamar_mandi": km_line,
                    "fasilitas": ", ".join(fasilitas),
                    "baris_harga": line.strip(),
                    "teks_asli": post,
                }

                hasil_data.append(data)
                nomor_id += 1

        if not ada_baris_harga:
            data = {
                "id": nomor_id,
                "nomor_postingan": nomor_post,
                "status": status,
                "judul": judul,
                "area": area,
                "tipe_kost": tipe_kost,
                "harga": None,
                "kamar_mandi": ambil_status_kamar_mandi(post),
                "fasilitas": ", ".join(fasilitas),
                "baris_harga": None,
                "teks_asli": post,
            }

            hasil_data.append(data)
            nomor_id += 1

    df = pd.DataFrame(hasil_data)
    return df


def filter_kriteria(df):
    """
    Filter sesuai target:
    - Harga 650.000 sampai 750.000
    - Kamar mandi dalam
    - Tipe putra/campur
    - Area Maguwoharjo, Condongcatur, atau Jalan Kaliurang
    - Bukan postingan pencarian
    """
    area_target = ["Maguwoharjo", "Condongcatur", "Jalan Kaliurang"]

    hasil = df[
        (df["status"] == "Penawaran")
        & (df["harga"] >= 650)
        & (df["harga"] <= 750)
        & (df["area"].isin(area_target))
        & (df["kamar_mandi"] == "Dalam")
        & (df["tipe_kost"].str.contains("Putra|Campur", case=False, na=False))
    ]

    return hasil

def rapikan_excel(file_path):
    """
    Merapikan tampilan file Excel agar lebih mudah dibaca.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Bekukan baris header
    ws.freeze_panes = "A2"

    # Aktifkan filter
    ws.auto_filter.ref = ws.dimensions

    # Style header
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Lebar kolom manual
    lebar_kolom = {
        "A": 8,    # id
        "B": 16,   # nomor_postingan
        "C": 14,   # status
        "D": 35,   # judul
        "E": 18,   # area
        "F": 16,   # tipe_kost
        "G": 10,   # harga
        "H": 16,   # kamar_mandi
        "I": 35,   # fasilitas
        "J": 45,   # baris_harga
        "K": 80,   # teks_asli
    }

    for kolom, lebar in lebar_kolom.items():
        ws.column_dimensions[kolom].width = lebar

    # Wrap text semua cell
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(file_path)


def simpan_data(df, df_filter):
    OUTPUT_ALL_CSV.parent.mkdir(parents=True, exist_ok=True)

    df.to_csv(OUTPUT_ALL_CSV, index=False, encoding="utf-8-sig")
    df.to_excel(OUTPUT_ALL_EXCEL, index=False)

    df_filter.to_csv(OUTPUT_FILTER_CSV, index=False, encoding="utf-8-sig")
    df_filter.to_excel(OUTPUT_FILTER_EXCEL, index=False)

    # Rapikan tampilan Excel
    rapikan_excel(OUTPUT_ALL_EXCEL)
    rapikan_excel(OUTPUT_FILTER_EXCEL)

    print("\nData berhasil disimpan.")
    print(f"Semua data CSV    : {OUTPUT_ALL_CSV}")
    print(f"Semua data Excel  : {OUTPUT_ALL_EXCEL}")
    print(f"Filter CSV        : {OUTPUT_FILTER_CSV}")
    print(f"Filter Excel      : {OUTPUT_FILTER_EXCEL}")


if __name__ == "__main__":
    data_postingan = baca_data_mentah()

    df = buat_dataframe(data_postingan)
    df_filter = filter_kriteria(df)

    print("SEMUA DATA:")
    print("=" * 100)
    print(df[["id", "status", "area", "tipe_kost", "harga", "kamar_mandi", "baris_harga"]])

    print("\nDATA SESUAI KRITERIA:")
    print("=" * 100)

    if df_filter.empty:
        print("Belum ada data yang sesuai kriteria.")
    else:
        print(df_filter[["id", "area", "tipe_kost", "harga", "kamar_mandi", "baris_harga"]])

    simpan_data(df, df_filter)